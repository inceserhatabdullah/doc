#!/usr/bin/env python3

import openpyxl as xlsx

# Load file
xlsxFile = xlsx.load_workbook('./inventory.xlsx')

# grab the active worksheet
sheet1 = xlsxFile['Sheet1']

# calculate max_row count
sheet1_max_rows = sheet1.max_row + 1

sheet1_list = []
supplier_list = []

def getUniqueSupplierList():
    for row in range(2, sheet1_max_rows):
        supplier = sheet1.cell(row, 4).value
        if supplier in supplier_list:
            continue
        else:
            supplier_list.append(supplier)
    return supplier_list
    # return list(set(supplier_list))

getUniqueSupplierList()

def mapUniqueSupplierList():
    for index in range(len(supplier_list)):
        sheet1_list.append({
            'name': supplier_list[index],
            'price': 0,
            'count': 0,
        })
    return sheet1_list

mapUniqueSupplierList()
    
for row in range(2, sheet1_max_rows):
    product_number = sheet1.cell(row, 1).value
    inventory = sheet1.cell(row, 2).value
    price = sheet1.cell(row, 3).value
    supplier = sheet1.cell(row, 4).value
   
    for index in range(len(supplier_list)):
        if supplier in sheet1_list[index]['name']:
           sheet1_list[index]['price'] += price * inventory
           sheet1_list[index]['count'] += 1
           break


# print(f'Supplier List => {sheet1_list}')